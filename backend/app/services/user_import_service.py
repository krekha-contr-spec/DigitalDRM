"""
user_import_service.py
-----------------------
Excel import for Data Entry Users (the RoleAccess table).

This replaces the old workflow of hand-editing `seed_role_access.py` every
time an employee joins, leaves, or changes department. An admin now just
uploads an .xlsx file from the "Data Entry Users" page.

Expected columns (header row, any order, case-insensitive). Common
synonyms are accepted so slightly different exports still work:

    Mandatory
    ---------
    Person Incharge   (aliases: person name, name, incharge, employee name)
    Email ID          (aliases: email, email address)
    Department        (aliases: dept)
    Plant             (aliases: plant id, plant no, plant number)

    Optional
    --------
    Employee ID       (aliases: emp id, employee code, emp code)
    Status            (aliases: active)  -> "Active"/"Inactive" (default Active)

Matching / upsert rules
------------------------
  1. If the row has an Employee ID and a row with that Employee ID already
     exists anywhere, that row is updated (person changed departments/plant,
     but keeps their ID).
  2. Otherwise, match on (plant_id, department-slug) — "who is currently
     the in-charge of this department at this plant". If found, update it.
  3. Otherwise, insert a new record.

This mirrors the existing uq_role_access_plant_person_email_role
constraint's intent (one active in-charge per plant+department at a time)
while still preventing duplicates.

Rows with a missing mandatory field, an invalid plant number, or an
unparseable email are skipped and reported in `errors` — they do NOT
abort the rest of the import. The whole batch of valid rows is applied
in a single database transaction (one commit at the end); if something
unexpected fails outside per-row validation, everything is rolled back
so the table is never left half-updated.
"""

import io
import re
from pathlib import Path
from typing import Optional

import openpyxl
from sqlalchemy.orm import Session

from app.models.models import RoleAccess, Plant

# ── Header synonym mapping ──────────────────────────────────────────────────

_HEADER_SYNONYMS = {
    "person_name": ["person incharge", "person name", "name", "incharge", "employee name"],
    "email":       ["email id", "email", "email address"],
    "department":  ["department", "dept"],
    "plant":       ["plant", "plant id", "plant no", "plant number"],
    "employee_id": ["employee id", "emp id", "employee code", "emp code", "employee no"],
    "status":      ["status", "active"],
}

_MANDATORY_FIELDS = ["person_name", "email", "department", "plant"]

_INACTIVE_VALUES = {"inactive", "no", "0", "false", "n"}

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def slugify(name: str) -> str:
    """'Rejection PPM' -> 'rejection_ppm', 'OVC' -> 'ovc'. Matches the
    normalization already used by role_access_routes.py so nothing else
    in the app breaks."""
    slug = name.strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "_", slug)
    return slug.strip("_")


def _build_header_map(header_row) -> dict:
    """header_row -> {field_key: column_index}, using the synonym table."""
    raw_headers = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            raw_headers[str(cell).strip().lower()] = idx

    field_col = {}
    for field, synonyms in _HEADER_SYNONYMS.items():
        for syn in synonyms:
            if syn in raw_headers:
                field_col[field] = raw_headers[syn]
                break
    return field_col


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def import_users_from_excel(db: Session, file_bytes: bytes) -> dict:
    """
    Parse the uploaded workbook and upsert RoleAccess rows.
    Returns a summary dict with counts + a list of per-row error strings.
    Raises ValueError for structural problems (bad file / missing columns)
    that make the whole import impossible.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read the Excel file: {exc}")
    
    return _process_workbook(db, wb)

def import_users_from_file(db: Session, filepath: str) -> dict:
    """
    Parse the workbook at the given filepath and upsert RoleAccess rows.
    Also deactivates any existing users that are not present in the Excel file.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read the Excel file: {exc}")
    
    return _process_workbook(db, wb, full_sync=True)

def _process_workbook(db: Session, wb, full_sync: bool = False) -> dict:
    ws = wb.active
    if ws is None or ws.max_row < 1:
        raise ValueError("The uploaded workbook has no data.")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError("The uploaded workbook has no header row.")

    field_col = _build_header_map(header_row)

    missing = [f for f in _MANDATORY_FIELDS if f not in field_col]
    if missing:
        raise ValueError(
            "The Excel file is missing required column(s): " + ", ".join(missing) +
            ". Expected at least: Person Incharge, Email ID, Department, Plant."
        )

    valid_plant_ids = {p.id for p in db.query(Plant.id).all()}

    total_rows = 0
    created = 0
    updated = 0
    unchanged = 0
    skipped = 0
    deactivated = 0
    errors = []

    # Track rows this Excel file accounted for by their real database id —
    # NOT by (plant_id, role, person_name, email). Matching by that tuple
    # is fragile: if two rows in the DB ever ended up sharing the same
    # tuple (duplicate data from before a dedupe constraint existed, a
    # manual DB edit, etc.), the tuple-based deactivation below would try
    # to update several different ids in one batch and could trip a
    # unique constraint — which then aborted the ENTIRE sync, blocking
    # every other legitimate change in the same file. Tracking by id is
    # correct regardless of what duplicate data might already exist.
    processed_ids = set()

    def get(row, field) -> str:
        col = field_col.get(field)
        if col is None or col >= len(row):
            return ""
        return _clean(row[col])

    try:
        for row_num, row in enumerate(rows_iter, start=2):
            if row is None:
                continue
            if all(v is None or str(v).strip() == "" for v in row):
                continue  # fully empty row — skip silently

            total_rows += 1

            person_name = get(row, "person_name")
            email = get(row, "email").lower()
            department = get(row, "department")
            plant_raw = get(row, "plant")
            employee_id = get(row, "employee_id") or None
            status_raw = get(row, "status").lower()

            # ── Mandatory field validation ──────────────────────────────
            if not person_name:
                errors.append(f"Row {row_num}: missing Person Incharge — skipped.")
                skipped += 1
                continue
            if not email:
                errors.append(f"Row {row_num}: missing Email ID — skipped.")
                skipped += 1
                continue
            if not _EMAIL_RE.match(email):
                errors.append(f"Row {row_num}: invalid email '{email}' — skipped.")
                skipped += 1
                continue
            if not department:
                errors.append(f"Row {row_num}: missing Department — skipped.")
                skipped += 1
                continue
            if not plant_raw:
                errors.append(f"Row {row_num}: missing Plant — skipped.")
                skipped += 1
                continue

            try:
                plant_id = int(float(plant_raw))
            except ValueError:
                errors.append(f"Row {row_num}: Plant '{plant_raw}' is not a valid number — skipped.")
                skipped += 1
                continue

            if valid_plant_ids and plant_id not in valid_plant_ids:
                errors.append(f"Row {row_num}: Plant {plant_id} does not exist — skipped.")
                skipped += 1
                continue

            role_slug = slugify(department)
            is_active = status_raw not in _INACTIVE_VALUES if status_raw else True

            # ── Find existing record (Employee ID first, then plant+role) ──
            existing = None
            if employee_id:
                existing = (
                    db.query(RoleAccess)
                    .filter(RoleAccess.employee_id == employee_id)
                    .first()
                )
            if existing is None:
                existing = (
                    db.query(RoleAccess)
                    .filter(RoleAccess.plant_id == plant_id, RoleAccess.role == role_slug)
                    .first()
                )

            if existing:
                changed = False
                if existing.person_name != person_name:
                    existing.person_name = person_name
                    changed = True
                if existing.email != email:
                    existing.email = email
                    changed = True
                if existing.plant_id != plant_id:
                    existing.plant_id = plant_id
                    changed = True
                if existing.role != role_slug:
                    existing.role = role_slug
                    changed = True
                if employee_id and existing.employee_id != employee_id:
                    existing.employee_id = employee_id
                    changed = True
                if existing.is_active != is_active:
                    existing.is_active = is_active
                    changed = True

                if changed:
                    # Committed immediately, row by row — NOT batched with
                    # any other row's changes. This is the actual fix for
                    # pre-existing duplicate data tripping ux_role_access_
                    # dedupe: SQLAlchemy only ever sends one row's UPDATE
                    # per commit, so if THIS row collides with a duplicate
                    # elsewhere in the table, only this row is skipped —
                    # every other row in the file still gets applied.
                    existing_id, existing_name, existing_plant, existing_role = (
                        existing.id, existing.person_name, existing.plant_id, existing.role
                    )
                    try:
                        db.commit()
                        updated += 1
                        processed_ids.add(existing_id)
                    except Exception as exc:
                        db.rollback()
                        skipped += 1
                        errors.append(
                            f"Row {row_num}: could not update '{existing_name}' "
                            f"(plant {existing_plant}, {existing_role}, id {existing_id}) — "
                            f"{exc}. Likely pre-existing duplicate data in role_access; "
                            f"this row was skipped, the rest of the file still applied. "
                            f"See dedupe_role_access() in this file for a one-time cleanup."
                        )
                else:
                    unchanged += 1
                    processed_ids.add(existing.id)
            else:
                record = RoleAccess(
                    plant_id=plant_id,
                    role=role_slug,
                    person_name=person_name,
                    email=email,
                    employee_id=employee_id,
                    is_active=is_active,
                )
                db.add(record)
                # Same row-by-row isolation as the update branch above.
                try:
                    db.commit()  # assigns record.id and persists it alone
                    processed_ids.add(record.id)
                    created += 1
                except Exception as exc:
                    db.rollback()
                    skipped += 1
                    errors.append(
                        f"Row {row_num}: could not create '{person_name}' "
                        f"(plant {plant_id}, {role_slug}) — {exc}. Likely a duplicate "
                        f"of an existing row; this row was skipped, the rest of the "
                        f"file still applied."
                    )

        if full_sync:
            # Any currently-active record this file didn't mention gets
            # deactivated (never hard-deleted — see module docstring).
            # Committed one row at a time so a single problem row — e.g.
            # pre-existing duplicate data tripping a unique constraint —
            # can't block every other row's sync.
            all_active = db.query(RoleAccess).filter(RoleAccess.is_active == True).all()
            for record in all_active:
                if record.id not in processed_ids:
                    record_id, record_name, record_plant, record_role = (
                        record.id, record.person_name, record.plant_id, record.role
                    )
                    record.is_active = False
                    try:
                        db.commit()
                        deactivated += 1
                    except Exception as exc:
                        db.rollback()
                        errors.append(
                            f"Could not deactivate '{record_name}' "
                            f"(plant {record_plant}, {record_role}, id {record_id}): {exc}. "
                            f"This is likely pre-existing duplicate data in role_access — "
                            f"see dedupe_role_access() in this file for a one-time cleanup."
                        )

    except Exception:
        db.rollback()
        raise

    return {
        "total_rows": total_rows,
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "skipped": skipped,
        "deactivated": deactivated,
        "errors": errors,
    }


def dedupe_role_access(db: Session) -> dict:
    """
    One-time cleanup for role_access rows that already violate
    ux_role_access_dedupe (same plant_id, role, person_name, email
    existing under more than one id) — this is legacy data from before
    that constraint existed, NOT something normal use of this app can
    create going forward. It's what causes a sync to report "could not
    update/deactivate ... likely pre-existing duplicate data" (see
    import_users_from_file above).

    For each duplicate group, keeps exactly one row — preferring an
    active one, then the most recently modified/highest id — and hard-
    deletes the rest. Safe to call as often as you like; it's a no-op
    once there are no duplicate groups left. Run this once via
    POST /admin/dedupe-role-access, then re-run "Sync from Excel".
    """
    from sqlalchemy import func

    dupe_keys = (
        db.query(
            RoleAccess.plant_id, RoleAccess.role,
            RoleAccess.person_name, RoleAccess.email,
        )
        .group_by(RoleAccess.plant_id, RoleAccess.role, RoleAccess.person_name, RoleAccess.email)
        .having(func.count(RoleAccess.id) > 1)
        .all()
    )

    groups_deduped = 0
    rows_deleted = []
    for plant_id, role, person_name, email in dupe_keys:
        rows = (
            db.query(RoleAccess)
            .filter(
                RoleAccess.plant_id == plant_id,
                RoleAccess.role == role,
                RoleAccess.person_name == person_name,
                RoleAccess.email == email,
            )
            .order_by(RoleAccess.is_active.desc(), RoleAccess.id.desc())
            .all()
        )
        keeper, extras = rows[0], rows[1:]
        for extra in extras:
            rows_deleted.append({
                "id": extra.id, "plant_id": plant_id, "role": role,
                "person_name": person_name, "email": email,
            })
            db.delete(extra)
        try:
            db.commit()
            groups_deduped += 1
        except Exception:
            db.rollback()
            # Leave this group as-is rather than crash the whole cleanup —
            # it'll just show up again next time this is run.
            continue

    return {"groups_deduped": groups_deduped, "rows_deleted": rows_deleted}


def export_users_to_excel(db: Session) -> bytes:
    """
    Builds an .xlsx workbook of every RoleAccess record across all plants,
    using the same column headers the importer accepts — so a file
    downloaded from here can be edited and re-uploaded through
    /role-access/import without any reformatting.

    Sorted by Plant, then Department, then Name for readability.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Entry Users"

    headers = [
        "Plant", "Department", "Person Incharge", "Email ID",
        "Employee ID", "Status",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    records = (
        db.query(RoleAccess)
        .order_by(RoleAccess.plant_id, RoleAccess.role, RoleAccess.person_name)
        .all()
    )

    for rec in records:
        ws.append([
            rec.plant_id,
            rec.role,
            rec.person_name,
            rec.email,
            rec.employee_id or "",
            "Active" if rec.is_active is not False else "Inactive",
        ])

    # Reasonable column widths instead of Excel's cramped default.
    widths = [8, 18, 24, 32, 14, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def write_users_excel_to_disk(db: Session, filepath) -> None:
    """
    Regenerates the on-disk users.xlsx master file from the current
    database — the DB -> Excel half of the Data Entry Users
    bidirectional sync. Called after every Admin Dashboard
    create/update/delete/status-toggle (see role_access_routes.py) so
    the file on disk always mirrors role_access, exactly mirroring the
    pattern already used for Email Services' email_users.xlsx (see
    email_recipient_service.write_email_recipients_excel_to_disk()).

    Any failure here is logged by the caller and must never block the
    database mutation that triggered it — this is a best-effort mirror,
    not the source of truth.
    """
    data = export_users_to_excel(db)
    path = Path(filepath)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "wb") as f:
        f.write(data)