"""
user_admin_service.py
----------------------
Business logic for the Admin's "Data Entry Users Management" module.
Handles create/update/delete/activate-deactivate and bulk import from
Excel for Data Entry User accounts (role = 'plant' / 'president', etc).

Admin accounts are never touched by any function here — every query is
explicitly scoped to exclude role == 'admin'.
"""

from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.workbook import Workbook
from sqlalchemy.orm import Session

from app.models.models import User
from app.services.auth import hash_password

ADMIN_ROLE = "admin"


def _read_header_map(ws) -> dict:
    headers = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = idx
    return headers


def list_data_entry_users(db: Session):
    return (
        db.query(User)
        .filter(User.role != ADMIN_ROLE)
        .order_by(User.username)
        .all()
    )


def get_data_entry_user(db: Session, user_id: int) -> Optional[User]:
    return (
        db.query(User)
        .filter(User.id == user_id, User.role != ADMIN_ROLE)
        .first()
    )


def create_data_entry_user(db: Session, username: str, password: str, role: str, plant_id: Optional[int]):
    if role == ADMIN_ROLE:
        raise ValueError("Cannot create another admin account through this module.")

    if db.query(User).filter(User.username == username).first():
        raise ValueError(f"Username '{username}' already exists.")

    user = User(
        username=username,
        password=hash_password(password),
        role=role,
        plant_id=plant_id,
        is_active=1,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return user


def update_data_entry_user(
    db: Session,
    user_id: int,
    password: Optional[str] = None,
    role: Optional[str] = None,
    plant_id: Optional[int] = None,
    is_active: Optional[bool] = None,
):
    user = get_data_entry_user(db, user_id)
    if not user:
        raise ValueError("Data entry user not found.")

    if role is not None:
        if role == ADMIN_ROLE:
            raise ValueError("Cannot promote a data entry user to admin through this module.")
        user.role = role

    if password:
        user.password = hash_password(password)

    if plant_id is not None:
        user.plant_id = plant_id

    if is_active is not None:
        user.is_active = 1 if is_active else 0

    db.commit()
    db.refresh(user)
    return user


def delete_data_entry_user(db: Session, user_id: int):
    user = get_data_entry_user(db, user_id)
    if not user:
        raise ValueError("Data entry user not found.")
    db.delete(user)
    db.commit()


def set_active_status(db: Session, user_id: int, is_active: bool):
    user = get_data_entry_user(db, user_id)
    if not user:
        raise ValueError("Data entry user not found.")
    user.is_active = 1 if is_active else 0
    db.commit()
    db.refresh(user)
    return user


def import_users_from_excel(db: Session, filepath: str) -> dict:
    """
    Bulk-create/update Data Entry Users from an .xlsx file.

    Expected header row (any case/order):
      Username | Password | Role | Plant Id

    - `Password` is required only for NEW users. If omitted for an
      existing username, the password is left unchanged.
    - Rows with role == 'admin' are rejected/skipped for safety.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb: Workbook = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook contains no worksheet.")

    headers = _read_header_map(ws)
    required = ["username", "role"]
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError("Missing columns: " + ", ".join(missing))

    col_username = headers["username"]
    col_password = headers.get("password")
    col_role = headers["role"]
    col_plant = headers.get("plant id", headers.get("plant_id"))

    created, updated, skipped = 0, 0, 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        try:
            username_raw = row[col_username]
            role_raw = row[col_role]

            if username_raw in (None, "") or role_raw in (None, ""):
                skipped += 1
                continue

            username = str(username_raw).strip()
            role = str(role_raw).strip().lower()

            if role == ADMIN_ROLE:
                errors.append(f"Row {row_num}: skipped — cannot import an admin account.")
                skipped += 1
                continue

            password_raw = row[col_password] if col_password is not None and col_password < len(row) else None
            plant_id_raw = row[col_plant] if col_plant is not None and col_plant < len(row) else None
            plant_id = int(plant_id_raw) if plant_id_raw not in (None, "") else None

            existing = db.query(User).filter(User.username == username).first()

            if existing:
                if existing.role == ADMIN_ROLE:
                    errors.append(f"Row {row_num}: skipped — '{username}' is an admin account.")
                    skipped += 1
                    continue
                existing.role = role
                existing.plant_id = plant_id
                if password_raw not in (None, ""):
                    existing.password = hash_password(str(password_raw))
                updated += 1
            else:
                if password_raw in (None, ""):
                    errors.append(f"Row {row_num}: skipped — password required for new user '{username}'.")
                    skipped += 1
                    continue
                db.add(
                    User(
                        username=username,
                        password=hash_password(str(password_raw)),
                        role=role,
                        plant_id=plant_id,
                        is_active=1,
                    )
                )
                created += 1

        except Exception as exc:
            errors.append(f"Row {row_num}: {exc}")
            skipped += 1

    db.commit()

    return {
        "created": created,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }
