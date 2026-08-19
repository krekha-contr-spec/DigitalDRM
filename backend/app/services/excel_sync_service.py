"""
excel_sync_service.py
----------------------
Reads the User Master and Department Master Excel files and syncs their
contents into the database. This replaces manually editing
seed_role_access.py every time an employee joins/leaves.

Expected file: data/user_master.xlsx
Columns (header row, any case/order): Plant | Department | Person Name | Email

Expected file: data/department_master.xlsx
Columns (header row): Department

How matching works
-------------------
Each department name (e.g. "Rejection PPM") is converted to a machine
slug (e.g. "rejection_ppm") via slugify(). That slug is:
  - stored in the Department master table, and
  - used as RoleAccess.role (kept identical to today's values like
    "production", "ovc", "product_value" so nothing else in the app
    breaks).

A row in user_master.xlsx is matched to an existing RoleAccess record by
(plant_id, role_slug) — i.e. "who currently holds this department in this
plant". If the Excel row's name/email differs from what's in the DB, the
existing record is UPDATED (handles "Production In-charge changed from
X to Y"). If no record exists yet, a new one is INSERTED. Nothing is
ever duplicated because the (plant_id, role_slug) pair is the identity.
"""

import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.models import Department, RoleAccess


def slugify(name: str) -> str:
    """'Rejection PPM' -> 'rejection_ppm', 'OVC' -> 'ovc'."""
    slug = name.strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "_", slug)
    return slug.strip("_")


def _read_header_map(ws: Worksheet) -> dict[str, int]:
    """Map lowercase header text -> 0-based column index, from row 1."""
    headers = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value is not None:
            headers[str(cell.value).strip().lower()] = idx
    return headers


def _get_or_create_department(db: Session, display_name: str) -> Department:
    slug = slugify(display_name)
    dept = db.query(Department).filter(Department.slug == slug).first()
    if dept:
        # Keep the nicer/most-recent display name and make sure it's active.
        changed = False
        if dept.name != display_name:
            dept.name = display_name
            changed = True
        if not dept.is_active:
            dept.is_active = True
            changed = True
        if changed:
            db.flush()
        return dept

    dept = Department(name=display_name, slug=slug, is_active=True)
    db.add(dept)
    db.flush()
    return dept


def sync_departments_from_excel(db: Session, filepath: str) -> dict:
    """Sync department_master.xlsx -> Department table."""

    path = Path(filepath)

    if not path.exists():
        raise FileNotFoundError(f"Department master file not found: {filepath}")

    wb: Workbook = openpyxl.load_workbook(str(path), data_only=True)

    ws = wb.active
    if ws is None:
        raise ValueError("Department workbook contains no worksheet.")

    headers = _read_header_map(ws)

    if "department" not in headers:
        raise ValueError(
            "department_master.xlsx must contain a 'Department' column."
        )

    dept_col = headers["department"]

    created = 0
    updated = 0
    seen = 0

    for row in ws.iter_rows(min_row=2, values_only=True):

        if row is None:
            continue

        if dept_col >= len(row):
            continue

        if row[dept_col] in (None, ""):
            continue

        name = str(row[dept_col]).strip()

        if not name:
            continue

        seen += 1

        existed = (
            db.query(Department)
            .filter(Department.slug == slugify(name))
            .first()
            is not None
        )

        _get_or_create_department(db, name)

        if existed:
            updated += 1
        else:
            created += 1

    db.commit()

    return {
        "rows_seen": seen,
        "created": created,
        "updated_or_unchanged": updated,
        "total_departments_now": db.query(Department).count(),
    }
def sync_users_from_excel(db: Session, filepath: str) -> dict:
    """
    Sync user_master.xlsx -> RoleAccess table.
    """

    path = Path(filepath)

    if not path.exists():
        raise FileNotFoundError(f"User master file not found: {filepath}")

    wb: Workbook = openpyxl.load_workbook(str(path), data_only=True)

    ws = wb.active
    if ws is None:
        raise ValueError("User workbook contains no worksheet.")

    headers = _read_header_map(ws)

    required = ["plant", "department", "person name", "email"]

    missing = [x for x in required if x not in headers]

    if missing:
        raise ValueError(
            "Missing columns: "
            + ", ".join(missing)
        )

    col = {r: headers[r] for r in required}

    created = 0
    updated = 0
    unchanged = 0
    skipped = 0

    errors = []
    seen_keys = set()

    for row_num, row in enumerate(
        ws.iter_rows(min_row=2, values_only=True),
        start=2,
    ):

        if row is None:
            continue

        if all(v is None or str(v).strip() == "" for v in row):
            continue

        try:

            plant_raw = row[col["plant"]]
            dept_raw = row[col["department"]]
            person_raw = row[col["person name"]]
            email_raw = row[col["email"]]

            if plant_raw in (None, ""):
                skipped += 1
                continue

            if dept_raw in (None, ""):
                skipped += 1
                continue

            if person_raw in (None, ""):
                skipped += 1
                continue

            plant_id = int(plant_raw)

            department_name = str(dept_raw).strip()

            person_name = str(person_raw).strip()

            email = ""

            if email_raw not in (None, ""):
                email = str(email_raw).strip()

            dept = _get_or_create_department(db, department_name)

            role_slug = dept.slug

            key = (plant_id, role_slug)

            seen_keys.add(key)

            existing = (
                db.query(RoleAccess)
                .filter(
                    RoleAccess.plant_id == plant_id,
                    RoleAccess.role == role_slug,
                )
                .first()
            )

            if existing:

                if (
                    existing.person_name != person_name
                    or existing.email != email
                ):
                    existing.person_name = person_name
                    existing.email = email
                    updated += 1
                else:
                    unchanged += 1

            else:

                db.add(
                    RoleAccess(
                        plant_id=plant_id,
                        role=role_slug,
                        person_name=person_name,
                        email=email,
                    )
                )

                created += 1

        except Exception as exc:

            errors.append(f"Row {row_num}: {exc}")

            skipped += 1

    db.commit()

    return {
        "rows_processed": len(seen_keys),
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "skipped": skipped,
        "errors": errors,
    }

def sync_all(db: Session, dept_filepath: Optional[str], user_filepath: str) -> dict:
    """Convenience wrapper: sync departments (if file given) then users."""
    dept_result = None
    if dept_filepath and Path(dept_filepath).exists():
        dept_result = sync_departments_from_excel(db, dept_filepath)
    user_result = sync_users_from_excel(db, user_filepath)
    return {"departments": dept_result, "users": user_result}