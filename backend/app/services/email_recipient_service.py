"""
email_recipient_service.py
---------------------------
Central place that resolves WHO gets automated emails (missing-data
reminders, 3-level escalation, and Monthly/Quarterly/Yearly report
emails), backed entirely by the `email_recipients` table. This table —
managed exclusively through the Admin Dashboard's "Email Services" page
— is the single master source for every recipient address in the
system. Nothing in this codebase hardcodes a Staff Incharge, Plant Head,
President, or report recipient email address; whatever an admin sets
here is picked up automatically by reminder_service.py and
report_save_routes.py at send-time, with no code change or redeploy
required.

Where Staff Incharge recipients come from
-------------------------------------------
Every department's Staff Incharge is already defined elsewhere in the
project: the Admin Dashboard's "Data Entry Users" page (the role_access
table), which assigns exactly one person per (plant, department). Rather
than duplicating those addresses as a second hardcoded list here,
sync_staff_incharge_from_role_access() mirrors that existing assignment
into email_recipients automatically, for every plant. An admin can then
override any of these directly on the Email Services page — that edit
becomes the new master value and is never re-overwritten by the sync.

Resolution rules
-----------------
For a given (plant_id, department, recipient_type):
  1. Prefer an active row that matches this exact plant_id (and, for
     staff_incharge, this exact department).
  2. If none exists, fall back to an active row with plant_id = NULL
     (a "global" / all-plants default) for the same department/type.
  3. If still nothing, return an empty list — callers must treat an
     empty recipient list as "nothing to send" and log it, never crash.

department is only meaningful for recipient_type = "staff_incharge".
plant_head / president rows are looked up with department = NULL.

Scalable across plants
-----------------------
plant_id is a free integer, not a hardcoded list — an admin can add
recipients for Plant 1 through Plant 5 (or any future plant) purely by
adding rows through the Email Services page, and staff_incharge
recipients scale to every plant automatically via
sync_staff_incharge_from_role_access(). Nothing here assumes a fixed set
of plants.
"""

import io
import re
from pathlib import Path
from typing import List, Optional

import openpyxl
import openpyxl.styles
import openpyxl.utils

from sqlalchemy.orm import Session

from app.models.models import EmailRecipient

# The single on-disk master file for Email Services — the bidirectional
# counterpart to backend/data/users.xlsx for Data Entry Users. Every DB
# change (Admin Dashboard create/update/delete/status-toggle, or the
# role_access-derived auto-sync) rewrites this file so it always mirrors
# the database; a periodic scheduler job (see app/scheduler.py) reads it
# back so edits made directly to the file also flow into the database —
# see write_email_recipients_excel_to_disk() / import_email_recipients_from_file().
EMAIL_USERS_XLSX_PATH = Path(__file__).resolve().parent.parent.parent / "data" / "email_users.xlsx"

RECIPIENT_TYPES = [
    "staff_incharge",
    "plant_head",
    "president",
    "admin",
    "overall_summary_recipient",
    "daily_report_recipient",
]

_TYPE_LABELS = {
    "staff_incharge": "Staff Incharge",
    "plant_head": "Plant Head",
    "president": "President",
    "admin": "Admin (User Approval Notifications)",
    "overall_summary_recipient": "Overall Summary Recipient",
    "daily_report_recipient": "Daily Report Recipient (Staff Incharge)",
}

# Accepted spellings/aliases in an uploaded Recipient Type column, mapped
# to the internal recipient_type key. Keys here are already lowercased.
# Note: "report_recipient" (Combined Report Recipient) was removed —
# Monthly/Quarterly/Yearly reports now go to "president" like every
# other plant-wide report; there is only one President per plant (or
# globally for All Plants), not a separate combined-report role. Any
# spelling that used to mean "report recipient" now maps to "president"
# so old uploaded sheets/exports keep working instead of erroring out.
_TYPE_ALIASES = {
    "staff incharge": "staff_incharge",
    "staff in-charge": "staff_incharge",
    "staff_incharge": "staff_incharge",
    "plant head": "plant_head",
    "plant_head": "plant_head",
    "president": "president",
    "admin": "admin",
    "combined report recipient": "president",
    "report recipient": "president",
    "report_recipient": "president",
    "overall summary recipient": "overall_summary_recipient",
    "overall summary recipient (president dashboard)": "overall_summary_recipient",
    "overall_summary_recipient": "overall_summary_recipient",
    "daily report recipient": "daily_report_recipient",
    "daily report recipient (staff incharge)": "daily_report_recipient",
    "daily_report_recipient": "daily_report_recipient",
}

_INACTIVE_VALUES = {"inactive", "no", "0", "false", "n"}
_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


# ── Resolution (used by reminder_service.py / report_save_routes.py) ───────

def get_recipients(
    db: Session,
    recipient_type: str,
    department: Optional[str] = None,
    plant_id: Optional[int] = None,
) -> List[str]:
    """
    Returns the list of active email addresses configured for this
    (plant_id, department, recipient_type). Falls back to a global
    (plant_id IS NULL) row if no plant-specific one exists. Returns []
    if nothing is configured — callers must handle that gracefully.

    Department matching is normalized (via _dept_slugify — same
    normalization Excel import already uses) rather than a raw exact
    string match. This is deliberate: department values can end up
    stored a few different ways over time — the canonical slug
    (reminder_service.py's dept["slug"], e.g. "ovc",
    "rejection_ppm"), a human label typed by an admin ("OVC Elements",
    "Rejection PPM"), or something with stray whitespace/casing from a
    manual edit or an older Excel upload — and a raw `==` comparison
    would silently return zero recipients (which reminder_service.py
    then logs as "No recipients configured" and skips the email
    entirely) even though a recipient genuinely IS configured, just
    spelled slightly differently. Normalizing both sides here means a
    Level 1 reminder finds its Staff Incharge no matter which spelling
    was used when the row was added.
    """
    query = db.query(EmailRecipient).filter(
        EmailRecipient.recipient_type == recipient_type,
        EmailRecipient.is_active == True,  # noqa: E712
    )

    target_dept = _dept_slugify(department) if department else None

    def _matches_department(rows):
        if target_dept is None:
            return [r for r in rows if not r.department]
        return [r for r in rows if r.department and _dept_slugify(r.department) == target_dept]

    if plant_id is not None:
        plant_specific = _matches_department(query.filter(EmailRecipient.plant_id == plant_id).all())
        if plant_specific:
            return [r.email for r in plant_specific]

    global_rows = _matches_department(query.filter(EmailRecipient.plant_id.is_(None)).all())
    return [r.email for r in global_rows]


# ── Admin CRUD (used by email_recipient_routes.py) ──────────────────────────

def list_recipients(db: Session):
    return (
        db.query(EmailRecipient)
        .order_by(EmailRecipient.plant_id, EmailRecipient.recipient_type, EmailRecipient.department)
        .all()
    )


def create_recipient(
    db: Session,
    recipient_type: str,
    email: str,
    department: Optional[str] = None,
    plant_id: Optional[int] = None,
    name: Optional[str] = None,
    is_active: bool = True,
) -> EmailRecipient:
    if recipient_type not in RECIPIENT_TYPES:
        raise ValueError(f"Invalid recipient_type '{recipient_type}'.")
    if recipient_type == "staff_incharge" and not department:
        raise ValueError("Staff Incharge recipients require a department.")
    if recipient_type in ("plant_head", "president") and department:
        raise ValueError(f"{recipient_type} recipients cannot have a department.")

    record = EmailRecipient(
        plant_id=plant_id,
        department=_dept_slugify(department) if department else None,
        recipient_type=recipient_type,
        name=name,
        email=email.strip().lower(),
        is_active=is_active,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    return record


def update_recipient(
    db: Session,
    recipient_id: int,
    recipient_type: Optional[str] = None,
    email: Optional[str] = None,
    department: Optional[str] = None,
    plant_id: Optional[int] = None,
    name: Optional[str] = None,
    is_active: Optional[bool] = None,
    clear_plant_id: bool = False,
    clear_department: bool = False,
) -> EmailRecipient:
    record = db.query(EmailRecipient).filter(EmailRecipient.id == recipient_id).first()
    if not record:
        raise ValueError("Email recipient not found.")

    if recipient_type is not None:
        record.recipient_type = recipient_type
    if email is not None:
        record.email = email.strip().lower()
    if name is not None:
        record.name = name
    if is_active is not None:
        record.is_active = is_active
    if plant_id is not None:
        record.plant_id = plant_id
    elif clear_plant_id:
        record.plant_id = None
    if department is not None:
        record.department = _dept_slugify(department)
    elif clear_department:
        record.department = None

    db.commit()
    db.refresh(record)
    return record


def delete_recipient(db: Session, recipient_id: int) -> None:
    record = db.query(EmailRecipient).filter(EmailRecipient.id == recipient_id).first()
    if not record:
        raise ValueError("Email recipient not found.")
    db.delete(record)
    db.commit()


def set_active_status(db: Session, recipient_id: int, is_active: bool) -> EmailRecipient:
    record = db.query(EmailRecipient).filter(EmailRecipient.id == recipient_id).first()
    if not record:
        raise ValueError("Email recipient not found.")
    record.is_active = is_active
    db.commit()
    db.refresh(record)
    return record


def export_email_recipients_to_excel(db: Session) -> bytes:
    """
    Builds an .xlsx workbook of every Email Services recipient (all
    plants), using the same column layout the importer accepts — so a
    file downloaded from here can be edited and re-uploaded through
    /admin/email-recipients/import without any reformatting. Mirrors
    export_users_to_excel() in user_import_service.py so the Email
    Services "Download Excel" action matches the Data Entry Users page
    exactly.

    Sorted by Plant, then Recipient Type, then Department for readability.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Services"

    headers = ["Plant ID", "Department", "Recipient Type", "Name", "Email ID", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    records = list_recipients(db)

    for rec in records:
        ws.append([
            rec.plant_id if rec.plant_id is not None else "",
            deptLabel := (_dept_label(rec.department) if rec.department else ""),
            _TYPE_LABELS.get(rec.recipient_type, rec.recipient_type),
            rec.name or "",
            rec.email,
            "Active" if rec.is_active is not False else "Inactive",
        ])

    widths = [10, 18, 22, 24, 32, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def write_email_recipients_excel_to_disk() -> None:
    """
    Regenerates backend/data/email_users.xlsx from the current database
    state and overwrites it on disk. This is what makes the DB -> Excel
    direction of the bidirectional sync automatic: every route that
    changes a recipient (create/update/delete/status toggle, or the
    role_access-derived auto-sync) calls this right after committing, so
    the on-disk file is never more than one request behind the database.

    Deliberately swallows its own errors (e.g. disk full, permissions) —
    a problem writing this convenience file must never fail the actual
    database change or break the API response the admin is waiting on.
    Uses its own short-lived session so it reflects only committed data.
    """
    import logging

    from app.database import SessionLocal

    logger = logging.getLogger("digitaldrm.email_recipient_service")

    db = SessionLocal()
    try:
        content = export_email_recipients_to_excel(db)
        EMAIL_USERS_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(EMAIL_USERS_XLSX_PATH, "wb") as f:
            f.write(content)
    except Exception as exc:
        logger.error("Could not write email_users.xlsx to disk: %s", exc)
    finally:
        db.close()


def generate_email_recipients_template() -> bytes:
    """
    Builds a blank .xlsx template with the correct headers and a few
    example rows covering every recipient type, for the Email Services
    page's "Download Template" action.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Email Services Template"

    headers = ["Plant ID", "Department", "Recipient Type", "Name", "Email ID", "Status"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    example_rows = [
        [5, "Production", "Staff Incharge", "Jane Doe", "jane.doe@example.com", "Active"],
        [5, "", "Plant Head", "John Smith", "john.smith@example.com", "Active"],
        ["", "", "President", "Ravi Kumar", "ravi.kumar@example.com", "Active"],
        ["", "", "Combined Report Recipient", "Report Desk", "reports@example.com", "Active"],
    ]
    for row in example_rows:
        ws.append(row)

    widths = [10, 18, 22, 24, 32, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Excel header parsing helpers (mirrors user_import_service.py) ──────────

_HEADER_SYNONYMS = {
    "plant_id": ["plant id", "plant", "plant no", "plant number"],
    "department": ["department", "dept"],
    "recipient_type": ["recipient type", "type"],
    "name": ["name"],
    "email": ["email id", "email", "email address"],
    "status": ["status", "active"],
}

# Department labels for the export — kept local (rather than imported
# from the frontend) so this file has no frontend dependency.
_DEPARTMENT_LABELS = {
    "production": "Production",
    "manpower": "Manpower",
    "ovc": "OVC",
    "despatch": "Despatch",
    "sales": "Sales",
    "rejection_ppm": "Rejection PPM",
    "product_value": "Product Value",
}


def _dept_label(slug: str) -> str:
    return _DEPARTMENT_LABELS.get(slug, slug)


def _dept_slugify(name: str) -> str:
    """'Rejection PPM' -> 'rejection_ppm', 'OVC' -> 'ovc'. Matches the
    normalization already used by user_import_service.slugify() so
    department values line up with role_access.role / reminder_service.py
    / report_save_routes.py without any extra mapping."""
    slug = name.strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "_", slug)
    return slug.strip("_")


def _build_header_map(header_row) -> dict:
    raw_headers = {}
    for idx, cell in enumerate(header_row):
        if cell is not None:
            raw_headers[str(cell).strip().lower()] = idx

    field_col = {}
    for field, synonyms in _HEADER_SYNONYMS.items():
        for syn in synonyms:
            if syn in raw_headers:
                field_col[field] = raw_headers[syn]
                break
    return field_col


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def import_email_recipients_from_excel(db: Session, file_bytes: bytes) -> dict:
    """
    Parse an uploaded workbook (Admin Dashboard > Email Services >
    Import) and upsert email_recipients rows. Returns a summary dict
    with counts + a list of per-row error strings. Raises ValueError for
    structural problems (bad file / missing columns) that make the whole
    import impossible.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read the Excel file: {exc}")
    return _process_email_workbook(db, wb)


def import_email_recipients_from_file(db: Session, filepath: str) -> dict:
    """
    Parse the workbook at the given filepath (the backend/data/
    email_users.xlsx master file, auto-synced on startup) and upsert
    email_recipients rows. Also deactivates any existing recipient that
    is not present in the file (full sync), matching the same behavior
    users.xlsx already has for Data Entry Users.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read the Excel file: {exc}")
    return _process_email_workbook(db, wb, full_sync=True)


def _process_email_workbook(db: Session, wb, full_sync: bool = False) -> dict:
    ws = wb.active
    if ws is None or ws.max_row < 1:
        raise ValueError("The uploaded workbook has no data.")

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError("The uploaded workbook has no header row.")

    field_col = _build_header_map(header_row)

    missing = [f for f in ("recipient_type", "email") if f not in field_col]
    if missing:
        raise ValueError(
            "The Excel file is missing required column(s): " + ", ".join(missing) +
            ". Expected at least: Recipient Type, Email ID (Plant ID, Department, "
            "Name, Status optional)."
        )

    try:
        from app.models.models import Plant
        valid_plant_ids = {p.id for p in db.query(Plant.id).all()}
    except Exception:
        valid_plant_ids = set()

    total_rows = 0
    created = 0
    updated = 0
    unchanged = 0
    skipped = 0
    deactivated = 0
    errors = []

    processed_keys = set()

    def get(row, field) -> str:
        col = field_col.get(field)
        if col is None or col >= len(row):
            return ""
        return _clean(row[col])

    try:
        for row_num, row in enumerate(rows_iter, start=2):
            if row is None:
                continue
            if all(v is None or str(v).strip() == "" for v in row):
                continue  # fully empty row — skip silently

            total_rows += 1

            plant_raw = get(row, "plant_id")
            department_raw = get(row, "department")
            type_raw = get(row, "recipient_type").lower()
            name = get(row, "name")
            email = get(row, "email").lower()
            status_raw = get(row, "status").lower()

            # ── Mandatory field validation ──────────────────────────────
            if not type_raw:
                errors.append(f"Row {row_num}: missing Recipient Type — skipped.")
                skipped += 1
                continue

            recipient_type = _TYPE_ALIASES.get(type_raw, type_raw.replace(" ", "_").replace("-", "_"))
            if recipient_type not in RECIPIENT_TYPES:
                errors.append(f"Row {row_num}: unknown Recipient Type '{get(row, 'recipient_type')}' — skipped.")
                skipped += 1
                continue

            if not email:
                errors.append(f"Row {row_num}: missing Email ID — skipped.")
                skipped += 1
                continue
            if not _EMAIL_RE.match(email):
                errors.append(f"Row {row_num}: invalid email '{email}' — skipped.")
                skipped += 1
                continue

            department = None
            if recipient_type == "staff_incharge":
                if not department_raw:
                    errors.append(f"Row {row_num}: Staff Incharge rows require a Department — skipped.")
                    skipped += 1
                    continue
                department = _dept_slugify(department_raw)
            elif department_raw:
                errors.append(
                    f"Row {row_num}: {_TYPE_LABELS.get(recipient_type, recipient_type)} rows must not "
                    f"have a Department — skipped."
                )
                skipped += 1
                continue

            plant_id = None
            if plant_raw:
                try:
                    plant_id = int(float(plant_raw))
                except ValueError:
                    errors.append(f"Row {row_num}: Plant '{plant_raw}' is not a valid number — skipped.")
                    skipped += 1
                    continue
                if valid_plant_ids and plant_id not in valid_plant_ids:
                    errors.append(f"Row {row_num}: Plant {plant_id} does not exist — skipped.")
                    skipped += 1
                    continue

            is_active = status_raw not in _INACTIVE_VALUES if status_raw else True

            # ── Find existing record: (plant_id, department, recipient_type) ──
            # is the exact slot get_recipients() resolves against, so
            # this is the natural upsert / de-dupe key.
            existing = (
                db.query(EmailRecipient)
                .filter(
                    EmailRecipient.plant_id == plant_id,
                    EmailRecipient.department == department,
                    EmailRecipient.recipient_type == recipient_type,
                )
                .first()
            )

            if existing:
                changed = False
                new_name = name or None
                if existing.name != new_name:
                    existing.name = new_name
                    changed = True
                if existing.email != email:
                    existing.email = email
                    changed = True
                if existing.is_active != is_active:
                    existing.is_active = is_active
                    changed = True
                if changed:
                    updated += 1
                else:
                    unchanged += 1
            else:
                db.add(EmailRecipient(
                    plant_id=plant_id,
                    department=department,
                    recipient_type=recipient_type,
                    name=name or None,
                    email=email,
                    is_active=is_active,
                ))
                created += 1

            processed_keys.add((plant_id, department, recipient_type))

        if full_sync:
            all_active = db.query(EmailRecipient).filter(EmailRecipient.is_active == True).all()  # noqa: E712
            for record in all_active:
                key = (record.plant_id, record.department, record.recipient_type)
                if key not in processed_keys:
                    record.is_active = False
                    deactivated += 1

        # Single transaction for the whole batch of valid rows.
        db.commit()

    except Exception:
        db.rollback()
        raise

    return {
        "total_rows": total_rows,
        "created": created,
        "updated": updated,
        "unchanged": unchanged,
        "skipped": skipped,
        "deactivated": deactivated,
        "errors": errors,
    }


# ── Email Services auto-population from Data Entry Users ───────────────────

def sync_staff_incharge_from_role_access(db: Session) -> dict:
    """
    Populates Email Services' Staff Incharge (Level 1) recipients
    directly from the Admin Dashboard's Data Entry Users (role_access)
    table — the department in-charge already assigned for every plant —
    instead of any hardcoded, department-wise email address. This is
    the ONLY source new Staff Incharge defaults come from; no email
    address for any department is written into this codebase.

    Scales to every plant automatically: any (plant_id, role) pair with
    an active Data Entry User gets a matching Staff Incharge recipient
    here, using that person's exact Name and Email — Plant 1 through
    Plant 5 today, and any future plant the same way, with no code
    change required.

    Duplicate-safe / non-destructive, and Email Services stays the
    master at send time: if a (plant_id, department, staff_incharge)
    slot in email_recipients already exists — from a previous run of
    this sync, or because an Admin added/edited/replaced it directly on
    the Email Services page — it is left completely untouched. From
    that point on, only an Admin's edit on the Email Services page
    changes what reminder_service.py / report_save_routes.py send to
    for that department.
    """
    from app.models.models import RoleAccess  # local import: avoids a circular import at module load time

    role_access_rows = (
        db.query(RoleAccess)
        .filter(
            RoleAccess.is_active == True,  # noqa: E712
            RoleAccess.plant_id.isnot(None),
            RoleAccess.role.isnot(None),
            RoleAccess.email.isnot(None),
        )
        .all()
    )

    created = 0
    for row in role_access_rows:
        existing = (
            db.query(EmailRecipient)
            .filter(
                EmailRecipient.plant_id == row.plant_id,
                EmailRecipient.department == row.role,
                EmailRecipient.recipient_type == "staff_incharge",
            )
            .first()
        )
        if existing:
            continue  # Email Services already has a recipient for this plant+department — never overwrite it

        db.add(EmailRecipient(
            plant_id=row.plant_id,
            department=row.role,
            recipient_type="staff_incharge",
            name=row.person_name,
            email=row.email,
            is_active=True,
        ))
        created += 1

    if created:
        db.commit()

    return {"created": created}


# ── Data Entry Users auto-population from Email Services ───────────────────

def sync_role_access_from_staff_incharge(db: Session) -> dict:
    """
    Populates the Admin Dashboard's "Data Entry Users" table (role_access)
    from this Email Services module's staff_incharge configuration, so
    every department in-charge already set up for automated notifications
    (Name, Email, Role/Department, Plant) also shows up as a Data Entry
    User automatically — no manual re-entry needed.

    Because role_access.role IS the department slug (production,
    manpower, ovc, rejection_ppm, product_value, despatch, sales — the
    same slugs used by email_recipients.department), each staff_incharge
    row maps directly onto exactly one role_access "department in-charge"
    slot for its plant.

    Duplicate-safe by design: a (plant_id, department) slot that already
    has a Data Entry User — whether created by a previous sync, the
    Excel importer, or an admin typing it in by hand — is left completely
    untouched. Only slots with nobody assigned yet are filled in, so
    admins can freely edit or replace any of these users afterward
    without their changes ever being overwritten by a later restart.

    This only affects the role_access table (who is registered as a
    plant's data-entry / department in-charge). It does NOT touch
    email_recipients, reminder_service.py, or report_save_routes.py —
    all existing email/notification routing logic and behavior is
    unchanged.
    """
    from app.models.models import RoleAccess  # local import: avoids a circular import at module load time

    staff_incharge_rows = (
        db.query(EmailRecipient)
        .filter(
            EmailRecipient.recipient_type == "staff_incharge",
            EmailRecipient.is_active == True,  # noqa: E712
            EmailRecipient.plant_id.isnot(None),
            EmailRecipient.department.isnot(None),
        )
        .all()
    )

    created = 0
    for row in staff_incharge_rows:
        already_assigned = (
            db.query(RoleAccess)
            .filter(
                RoleAccess.plant_id == row.plant_id,
                RoleAccess.role == row.department,
            )
            .first()
        )
        if already_assigned:
            continue  # this plant+department slot already has a Data Entry User — never overwrite it

        db.add(RoleAccess(
            plant_id=row.plant_id,
            role=row.department,
            person_name=row.name or row.email.split("@")[0],
            email=row.email,
            employee_id=None,
            is_active=True,
        ))
        created += 1

    if created:
        db.commit()

    return {"created": created}